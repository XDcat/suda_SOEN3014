# -*- coding:utf-8 -*-
'''
__author__ = 'XD'
__mtime__ = 2019/9/22
__project__ = 工业实践课程
Fix the Problem, Not the Blame.
'''

import os
import random
import re

import openpyxl


class Book():
    """
    可以进行一下几个操作，在不进行下列操作时，仅是一个普通的类。
    1. 增：在文件中增加自己的信息
        * 如果主键已经存在，则不能添加
        * 通过直接创建对象的，如果没有加入 id ，可以自动生成 id
    2. 删：删文件中删除自己的信息，如果提供 id 则默认删除选定 id 的
    3. 查：返回新的 Book 类
    4. 改：在文件中修改自己的信息，如果没有存在，增加记录
    """

    def __init__(self):
        pass

    def save(self):
        if not self.query(self.book_id):  # 在文件中没有该类的主键时，才可以插入
            if not self.book_id:
                self.book_id = self.find_max_id() + 1
            with open(self.file_path, "a+", encoding="utf-8") as f:
                f.write(self.__str__())
                f.write("\n")
                print("保存：%s" % self)
        else:
            print("=>错误：文件中已经存在主键 %s ，你可能需要更新操作" % self.book_id)

    def delete(self, book_id=None):
        if not book_id:
            # 删除指定 id 的
            book_id = self.book_id
        print("删除：book_id = %s" % self.query(book_id))

        with open(self.file_path, "r+", encoding="utf-8") as f:
            text = f.read()
        with open(self.file_path, "w+", encoding="utf-8") as f:
            f.write(re.sub("%s,.*?\n" % book_id, "", text))

    def delete_all(self):
        print("删除全部")
        open(self.file_path, "w+").close()

    def query_all(self):
        books = []
        with open(self.file_path, "r", encoding="utf-8") as f:
            for i in f.readlines():
                i = i.replace("\n", "")
                books.append(Book(*i.split(",")))
        return books

    def query(self, book_id):
        books = self.query_all()

        for i in books:
            if str(i.book_id) == str(book_id):
                return i

    def update(self):
        with open(self.file_path, "r+", encoding="utf-8") as f:
            text = f.read()
        with open(self.file_path, "w+", encoding="utf-8") as f:
            text = re.sub("%s,.*" % self.book_id, self.__str__(), text)
            f.write(text)

    def find_max_id(self):
        book_now = None
        with open(self.file_path, "r", encoding="utf-8") as f:
            for i in f.readlines():
                i = i.replace("\n", "")
                book_now = Book(*i.split(","))
        return book_now.book_id



    def print_file(self):
        with open(self.file_path, "r", encoding="utf-8") as f:
            print("\n==>文件中的结果（因为原文件有数据，所以条数可能不一致）")
            for i in f.readlines():
                print("\t", i, end="")


def get_many_books():
    """从 excel 中拿到数据"""
    wb = openpyxl.load_workbook(r".\data\book_list-计算机-机器学习-linux-android-数据库-互联网.xlsx")
    sheets = wb.sheetnames
    ws = wb[sheets[1]]
    title = [call.value for call in ws[1]]

    # 读取到对应的行
    ids = ws["A"][1:]
    names = ws["B"][1:]
    marks = ws["C"][1:]
    mark_peoples = ws["D"][1:]
    authors = ws["E"][1:]
    publishers = ws["F"][1:]

    # 提取有效数据
    ids = [i.value for i in ids]
    names = ["《%s》" % (i.value,) for i in names]
    marks = [i.value for i in marks]
    mark_peoples = [i.value for i in mark_peoples]
    authors = [i.value.replace("作者/译者：", "") for i in authors]
    publishers = [i.value.replace("出版信息：", "") for i in publishers]

    # 保存数据
    books_info = zip(ids, names, marks, mark_peoples, authors, publishers)
    books = []
    for i in books_info:
        i = [str(j).replace(",", "") for j in i]
        books.append(Book(*i))
    return books


if __name__ == '__main__':
    sample_books = get_many_books()
    # sample_books[0].delete_all()
    print("====== 1. 从 excel 中随机找到5本书的数据，将他们保存到文件中 =======")
    random_books = random.sample(sample_books, 5)
    # random_books = sample_books[:5]
    for i in random_books:
        i.save()
    random_books[0].print_file()  # 展示结果
    print()
    print()
    print()

    print("====== 2. 删除其中前 2 本 =======")
    for i in random_books[:2]:
        i.delete()
    random_books[0].print_file()  # 展示结果
    print()
    print()
    print()

    print("====== 3. 重复插入刚才的 5 本 =======")
    for i in random_books:
        i.save()
    random_books[0].print_file()  # 展示结果
    print()
    print()
    print()

    print("====== 4. 更新 id = %s 的 book =======" % random_books[-1].book_id)
    book = random_books[-2]
    print("改前：%s" % book)
    book.book_name = "《这是一个被改过的 book》"
    book.book_count_mark_people = 10000000000000000000000000000000000000
    book.update()
    print("改后：%s" % book)

    random_books[0].print_file()  # 展示结果
    print()
    print()
    print()

    print("====== 5. 查询 book =======")
    print("查询已经存在的书籍 id = %s" % random_books[-1].book_id)
    print(Book().query(random_books[-1].book_id))
    print("查询没有的数据 id = 1000")
    print(Book().query(100))
    print()
    print()
    print()

    print("====== 6. 查询 book，并当作对象取值 =======")
    print("查询已经存在的书籍 id = %s" % random_books[-1].book_id)
    book = Book().query(random_books[-1].book_id)

    print("book.book_id: ", book.book_id)
    print("book.book_name: ", book.book_name)
    print("book.book_mark: ", book.book_mark)
    print("book.book_count_mark_people: ", book.book_count_mark_people)
    print("book.book_author: ", book.book_author)
    print("book.book_publish: ", book.book_publish)
    print("\n====== Enter 键退出 ======")
    os.system("pause")
